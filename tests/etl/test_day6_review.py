"""Integration tests for Day 6: Parse-failure rejection + DQ-06 sector carve-out."""

from __future__ import annotations

import pandas as pd

from src.etl import load_dataset
from src.etl.database import get_connection, init_schema, load_dataframe
from src.etl.normalizers import YEAR_PARSE_ERROR


def _load_all_to(temp_db, raw_dir):
    """Helper: load all 12 datasets from raw_dir into temp_db, applying rejects."""
    from scripts.run_etl import _POST_PROCESS, LOAD_ORDER, _reject_critical_rows

    init_schema(temp_db)
    tables = {}
    rejects = []
    for name in LOAD_ORDER:
        df = load_dataset(name, data_dir=raw_dir)
        hook = _POST_PROCESS.get(name)
        if hook is not None:
            df = hook(df)
        before = len(df)
        df = _reject_critical_rows(df, name, rejects)
        assert len(df) <= before
        tables[name] = df
        load_dataframe(df, name, db_path=temp_db, deduplicate=True)
    return tables, rejects


def test_reject_critical_rows_filters_parse_error(tmp_path):
    """Rows with PARSE_ERROR year must be stripped before DB insertion."""

    from scripts.generate_data import generate_all

    raw = tmp_path / "raw"
    db = tmp_path / "nifty100.db"

    # 1. Generate all synthetic files into tmp raw/
    counts = generate_all(raw)
    assert counts["profitandloss"] > 0

    # 2. Inject a bad year directly into the generated P&L xlsx
    import openpyxl

    pnl_path = raw / "profitandloss.xlsx"
    wb = openpyxl.load_workbook(pnl_path)
    ws = wb["Profit & Loss"]
    # Core file: row 1 = metadata, row 2 = headers, data starts row 3.
    # Find year column header (row 2) and append a garbage row.
    header_row = [c.value for c in ws[2]]
    year_col_idx = header_row.index("year") + 1  # 1-indexed
    cid_col_idx = header_row.index("company_id") + 1
    new_row_idx = ws.max_row + 1
    ws.cell(row=new_row_idx, column=cid_col_idx, value="TCS")
    ws.cell(row=new_row_idx, column=year_col_idx, value="GARBAGE_YEAR")
    # Fill other columns with 0 so the row is otherwise valid-looking
    for ci in range(1, len(header_row) + 1):
        if ci in (cid_col_idx, year_col_idx):
            continue
        ws.cell(row=new_row_idx, column=ci, value=0.0)
    wb.save(pnl_path)

    # 3. Load through the pipeline
    _tables, rejects = _load_all_to(str(db), raw)

    # 4. Verify the bad row did NOT make it into the DB
    with get_connection(str(db)) as conn:
        cur = conn.execute("SELECT COUNT(*) FROM profitandloss WHERE year=?", (YEAR_PARSE_ERROR,))
        bad_in_db = cur.fetchone()[0]
        (total_pl,) = conn.execute("SELECT COUNT(*) FROM profitandloss").fetchone()

    assert bad_in_db == 0, "PARSE_ERROR rows must NOT be inserted into the DB"
    # One row should have been rejected
    assert len(rejects) >= 1
    reject_tables = {r["table"] for r in rejects}
    assert "profitandloss" in reject_tables
    # Total loaded should be one less than the original good count
    assert total_pl == counts["profitandloss"]


def test_reject_critical_rows_preserves_good_data(tmp_path):
    """When all years/tickers are valid, zero rows are rejected."""
    from scripts.generate_data import generate_all
    from scripts.run_etl import _reject_critical_rows

    raw = tmp_path / "raw"
    generate_all(raw)

    tables = {}
    from scripts.run_etl import _POST_PROCESS, LOAD_ORDER

    rejects = []
    for name in LOAD_ORDER:
        df = load_dataset(name, data_dir=raw)
        hook = _POST_PROCESS.get(name)
        if hook is not None:
            df = hook(df)
        df = _reject_critical_rows(df, name, rejects)
        tables[name] = df
    assert rejects == [], f"Expected no rejects, got: {rejects}"
    assert len(tables["companies"]) == 92
    assert len(tables["stock_prices"]) == 5520


def test_dq06_excludes_banks_when_sectors_present():
    """DQ-06 must not flag banks/NBFCs for sales<=0 when sectors are loaded."""
    from src.etl.validation import dq06_positive_sales

    pl = pd.DataFrame(
        {
            "company_id": ["HDFCBANK", "SBIN", "BAJFINANCE", "TCS", "INFY"],
            "year": ["2023-03"] * 5,
            "sales": [0.0, 0.0, 0.0, 100.0, 0.0],
        }
    )
    sectors = pd.DataFrame(
        {
            "company_id": ["HDFCBANK", "SBIN", "BAJFINANCE", "TCS", "INFY"],
            "broad_sector": [
                "Private Banks",
                "Public Banks",
                "Consumer Finance",
                "Information Technology",
                "Information Technology",
            ],
        }
    )
    failures = dq06_positive_sales({"profitandloss": pl, "sectors": sectors})
    flagged = {f.company_id for f in failures}
    # Only INFY is a non-bank with sales=0
    assert flagged == {"INFY"}, f"Expected {{INFY}}, got {flagged}"


def test_dq06_flags_all_zero_sales_when_no_sectors():
    """Back-compat: without sectors, every sales<=0 row is flagged."""
    from src.etl.validation import dq06_positive_sales

    pl = pd.DataFrame(
        {
            "company_id": ["HDFCBANK", "TCS", "INFY"],
            "year": ["2023-03"] * 3,
            "sales": [0.0, 100.0, 0.0],
        }
    )
    failures = dq06_positive_sales({"profitandloss": pl})
    flagged = {f.company_id for f in failures}
    assert flagged == {"HDFCBANK", "INFY"}


def test_parse_failures_csv_schema(tmp_path):
    """parse_failures.csv must have the expected columns when rejects exist."""

    from scripts.run_etl import _reject_critical_rows

    df = pd.DataFrame(
        {
            "id": [1, 2, 3],
            "company_id": ["TCS", "INFY", "TCS"],
            "year": ["2023-03", YEAR_PARSE_ERROR, "2022-03"],
            "sales": [100.0, 200.0, 90.0],
        }
    )
    rejects = []
    cleaned = _reject_critical_rows(df, "profitandloss", rejects)
    assert len(cleaned) == 2
    assert len(rejects) == 1
    assert rejects[0]["table"] == "profitandloss"
    assert rejects[0]["company_id"] == "INFY"
    assert rejects[0]["column"] == "year"
    assert "reason" in rejects[0] and rejects[0]["reason"]
